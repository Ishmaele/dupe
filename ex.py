from openpyxl import load_workbook
workbook = load_workbook(filename="work.xlsx")
sheet = workbook.active

listOfElems = []

for i in range(1, sheet.max_row+1):
    print(i, sheet.cell(row=i, column=1).value)
    listOfElems.append(sheet.cell(row=i, column=1).value)

print(listOfElems)