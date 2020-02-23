from openpyxl import load_workbook
from openpyxl import Workbook
from collections import Counter

workbook = load_workbook(filename="work.xlsx")
sheet = workbook.active

listOfElems = []

for i in range(1, sheet.max_row+1):
#   print(i, sheet.cell(row=i, column=1).value)
    listOfElems.append(sheet.cell(row=i, column=1).value)

# Create a dictionary of elements & their frequency count
dictOfElems = dict(Counter(listOfElems))
 
# Remove elements from dictionary whose value is 1, i.e. non duplicate items
dictOfElems = { key:value for key, value in dictOfElems.items() if value > 1}

workbookwr = Workbook()
sheetwr = workbookwr.active
sheetwr.column_dimensions['A'].width = 35
i=1

for key, value in dictOfElems.items():
    print('{:<30s}{:>10d}'.format(key,value))
    sheetwr.cell(row=i, column=1).value=key
    sheetwr.cell(row=i, column=2).value=value
    i=i+1

workbookwr.save(filename="result.xlsx") 